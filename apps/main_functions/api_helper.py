# -*- coding:utf-8 -*-
import re
from openpyxl import load_workbook
from io import BytesIO

from django.http import JsonResponse
from django.db.models import Q
from django.db import transaction

from apps.main_functions.functions import (
    object_fields,
    object_fields_types,
    object_foreign_keys,
)
from apps.main_functions.files import open_file, full_path
from apps.main_functions.model_helper import create_model_helper
from apps.main_functions.date_time import str_to_date

def ApiHelper(request,
              model_vars: dict,
              CUR_APP: str,
              reverse_params: dict = None,
              restrictions: list = None):
    """Апи-метод для получения всех данных
       :param request: HttpRequest
       :param model_vars: по какой модели отдаем данные
       :param CUR_APP: текущее приложение
       :param reverse_params: словарь для параметров в urls.py
       :param restrictions: аналог filters (массив Q() условий),
                            только уже из вьюхи, а по GET/POST
    """
    # Если нету метода, то не надо мучать жопу
    model = model_vars['model']
    with_get_folder = hasattr(model, 'get_folder')

    if not reverse_params:
        reverse_params = {}
    mh_vars = model_vars.copy()
    mh = create_model_helper(mh_vars, request, CUR_APP, reverse_params=reverse_params)
    # Принудительные права на просмотр
    mh.permissions['view'] = True

    # Независимо от метода
    params = request.GET if request.method == 'GET' else request.POST

    model_fields = object_fields(model())

    only_fields = []
    get_only_fields = params.get('only_fields')
    if get_only_fields:
        get_only_fields = [field.strip() for field in get_only_fields.split(',')]
        only_fields = [field for field in get_only_fields if field in model_fields]

    # Параметры фильтрации через filter__field = ''
    # filter__ означает, что мы хотим отфильтровать по какому то полю,
    # например, ?filter__id=1&filter__is_active=1
    filters = {k.replace('filter__', ''): v for k, v in params.items() if k.startswith('filter__')}
    if filters:
        mh.filter_add(Q(**filters))
    if restrictions:
        for restriction in restrictions:
            mh.filter_add(restriction)
    orders_by = {k.replace('order__', ''): v for k, v in params.items() if k.startswith('order__')}
    for order_by, direction in orders_by.items():
        if direction == 'asc':
            mh.order_by_add(order_by)
        elif direction == 'desc':
            mh.order_by_add('-%s' % order_by)

    context = mh.context
    rows = mh.standard_show(only_fields=only_fields)

    result = []
    for row in rows:
        item = object_fields(row, only_fields=only_fields)
        if not only_fields or 'folder' in only_fields:
            if with_get_folder:
                item['folder'] = row.get_folder()
        result.append(item)
    result = {'data': result,
              'last_page': mh.raw_paginator['total_pages'],
              'total_records': mh.raw_paginator['total_records'],
              'cur_page': mh.raw_paginator['cur_page'],
              'by': mh.raw_paginator['by'], }
    return JsonResponse(result, safe=False)

def XlsxHelper(request,
               model_vars: dict,
               CUR_APP: str,
               cond_fields: list = ['id'],
               reverse_params: dict = None):
    """Апи-метод для сохранения данных из excel-файла
       :param request: HttpRequest
       :param model_vars: по какой модели отдаем данные
       :param CUR_APP: текущее приложение
       :param cond_fields: поля, по которым определяем аналог для обновления
       :param reverse_params: словарь для параметров в urls.py по ним
                              хитрожопински обновляются foreign keys
                              через save_from_excel
    """
    result = {}
    mh_vars = model_vars.copy()

    if not reverse_params:
        reverse_params = {}

    mh = create_model_helper(mh_vars, request, CUR_APP, reverse_params=reverse_params)
    result['perms'] = mh.permissions

    action = request.POST.get('action')
    job = {'errors': ['Операция не поддерживается']}

    rows = []
    if action in ('save', 'drop'):
        count = request.POST.get('count')
        data = request.POST.get('data')
        if count.isdigit():
            count = int(count)
            template = object_fields(
                mh.model(),
                pass_fields=('created', 'updated', 'img')
            )
            if count > 0:
                search_names = [k for k in request.POST.keys() if k.startswith('data[0][')]

                # foreign_keys ключи, например, [0][key1][key2]
                rega_fk_fields = re.compile('\[([^\]]+?)\]', re.I+re.U+re.DOTALL)
                for search_name in search_names:
                    # Ну вот так коряво проверяем,
                    # что это со вложенным объектом,
                    # ну а хуля делать?
                    if search_name.count('][') > 1:
                        hier = rega_fk_fields.findall(search_name)[1:]
                        if hier:
                            # '[%s]' % ']['.join(['1','2','3']) = '[1][2][3]'
                            #names.append('[%s]' % ']['.join(hier))
                            template['__'.join(hier)] = None

                # тут проверяем именно наличие ключа,
                # а не значение
                names = [k for k in template.keys() if 'data[0][%s]' % k in search_names]

                # по шаблону модели обходим количество отправленных объектов
                rows = [{
                    k: request.POST.get('data[%s][%s]' % (i, k)) for k in names
                } for i in range(count)]

    # Предварительная подготовка данных
    # для просмотра перед загрузкой
    if action == 'import_xlsx':
        job = import_from_excel(mh.model,
                                request.FILES.get('file'))
        if not job['errors']:
            result['success'] = 'Файл обработан, проверьте данные и подтвердите операцию'
    elif action == 'save':
        if 'create' in mh.permissions and 'edit' in mh.permissions:
            job = save_from_excel(mh.model,
                                  rows,
                                  cond_fields=cond_fields,
                                  reverse_params=reverse_params)
        else:
            job = {'errors': ['Нужны права на создание и редактирование, чтобы выполнить загрузку файлом']}
        if not job['errors']:
            result['success'] = 'Добавлено %s, обновлено %s' % (job['created'], job['updated'])
    elif action == 'drop':
        if 'drop' in mh.permissions:
            job = drop_from_excel(mh.model, rows, cond_fields)
        else:
            job = {'errors': ['Нужны права на удаление, чтобы выполнить удаление файлом']}
        if not job['errors']:
            result['success'] = 'Удалено %s, Не найдено %s' % (job['dropped'], job['not_found'])
    result['resp'] = job
    if job['errors']:
        result['error'] = '<br>'.join(job['errors'])
    return JsonResponse(result, safe=False)

def open_wb(path, data_only: bool = True):
    """Загружаем эксельку
       :param path: путь до эксельки
       :param data_only: без формул - только значения
    """
    with open_file(path, 'rb') as excel_file:
        wb = load_workbook(BytesIO(excel_file.read()), data_only=data_only)
    # Посмотреть листы эксельки: logger.info(wb.sheetnames)
    return wb

def search_header(rows, symbols: str = '№'):
    """Поиск строки заголовков
       обход по генератору, необходимо помнить,
       что после вызова мы будем находиться на строк заголовка
       :param rows: генератор для строк
       :param symbols: символы с которых начинается заголовок таблицы
    """
    for row in rows:
        i = 0
        for cell in row:
            value = cell.value
            if value and symbols in str(value):
                return i, row
            i += 1
    return None, None

def accumulate_data(row, i, data):
    """Добавление ячейки в массив данных
       USAGE: data_arr = accumulate_data(row, i+1, result)

       где row - строка, где находимся генератором,
       i + 1 - номер ячейки по которой аккумулируем данные
       result - накопленные данные

       :param row: строка эксельки
       :param i: номер ячейки в строке эксельки
       :param data: массив данных
    """
    value = row[i].value
    if value:
        value = '%s' % value
        value = value.strip()
        if value and not value in data:
            data.append(value)
    return value

def import_from_excel(model,
                      excel_file,
                      required_names: list = None):
    """Импорт данных из excel
       :param model: модель, в которую импортируем данные
       :param excel_file: экселька, например, request.FILES.get("excel_file")
       :param required_names: список обязательных полей в шапке
    """
    if not required_names:
        required_names = []
    result = {'errors': []}
    if not excel_file:
        result['errors'].append('Не передан файл')
        return result
    wb = load_workbook(BytesIO(excel_file.read()))
    sheet = wb.active

    rows = list(sheet.rows)
    if not len(rows) > 0:
        result['errors'].append('Файл пустой')
        return result

    # ----------------
    # Обработка ошибок
    # ----------------
    errors = []
    names = [cell.value for cell in rows[0]]
    for name in required_names:
        if not name in names:
            result['errors'].append('обязательное поле %s' % name)

    if result['errors']:
        return result

    template = object_fields(model(), only_fields=names)
    indexes = {name: names.index(name) for name in names if name in template}

    # Добавить foreign keys в шапку
    with_fk = [name for name in names if '__' in name]
    if with_fk:
        fkeys = object_foreign_keys(model())
        for name in with_fk:
            wrong_key = False
            name_parts = name.split('__')
            for i in range(len(name_parts) - 1):
                if i == 0:
                    prev_fkeys = fkeys
                else:
                    prev_fkeys = object_foreign_keys(prev_fkeys[name_parts[i]])
                if not name_parts[i] in prev_fkeys:
                    wrong_key = True
            if not wrong_key:
                indexes[name] = names.index(name)

    def fill_fkey(item: dict, name: str, value: str):
        """Перезаполнить выходящее значение из строки в словарь
           :param item: словарь с текущими значениями
           :param name: Название поля
           :param value: Значение поля в виде словаря
        """
        parts = name.split('__')
        prev = item
        # Последняя вложенность - строкой (не словарем)
        for part in parts[:-1]:
            if not part in prev:
                prev[part] = {}
            prev = prev[part]
        prev[parts[-1]] = value

    data = []
    for row in rows:
        item = {}
        is_empty = True
        # В первой записи будут названия полей
        for name, ind in indexes.items():
            value = row[ind].value
            if value:
                is_empty = False
            # Будет дубль и key1__key2 и [key1][key2]
            item[name] = value

            # Спец обработка в словарь, если это foreign_key
            if name in with_fk:
                fill_fkey(item, name, value)
                continue

        if not is_empty:
            data.append(item)
    result['data'] = data[1:]
    return result

def save_from_excel(model,
                    data,
                    cond_fields: list = None,
                    reverse_params: dict = None):
    """Сохранение массива в базу
       :param model: модель, в которую импортируем данные
       :param data: массив с данными для импорта
       :param cond_fields: поля по которым ищем аналоги (обновляем)
       :param reverse_params: словарь параметров, которые
                              могут указывать на foreign keys

            # inject reverse params for build response
            # with correct foreign keys
            if reverse_params:
                item.update(reverse_params)
    """
    result = {'errors': []}
    if not data:
        result['errors'].append('Не переданы данные для сохранения')
        return result
    if not cond_fields:
        result['errors'].append('Не переданы условия для обновления')
        return result

    field_types = object_fields_types(model())
    fkeys = object_foreign_keys(model())

    # Небольшой хак на foreign key + reverse_params
    # Убираем _id, и,
    # если у нас есть такой foreign key,
    # пробуем записать
    fk_updates = {}
    if reverse_params:
        for rparam, rvalue in reverse_params.items():
            if not rparam.endswith('_id'):
                continue
            rparam_fk_name = rparam.replace('_id', '')
            if field_types.get(rparam_fk_name) == 'foreign_key':
                fk_updates[rparam] = rvalue

    create_tasks = []
    update_tasks = []
    cached_fk = {}

    for cond_field in cond_fields:
        if cond_field.count('__') == 1:
            del cond_fields[cond_fields.index(cond_field)]

    for item in data:
        cond = Q()
        # id обновлять - плохая затея
        for key in ('id', 'img', 'created', 'updated'):
            if key in item:
                del item[key]

        fordel = {}
        for k, v in item.items():
            if not k in field_types:
                fordel[k] = v
                continue
            field_type = field_types[k]
            # Некоторые типы полей пока пропускаем
            if field_type in ('foreign_key', ):
                continue
            if v:
                item[k] = v.strip()
            if field_type in ('date', 'datetime'):
                item[k] = str_to_date(item[k])
            elif field_type in ('int', ):
                try:
                    item[k] = int(item[k])
                except ValueError:
                    item[k] = None
            elif field_type in ('boolean', ):
                item[k] = True if item[k] in (1, '1') else False

        # Дропаем всякое говно, которое не в полях модели
        for k, v in fordel.items():
            del item[k]
            # Поддерживаем только одноуровневый fk
            # TODO: адаптировать под key1__key2__... (больше 1) вложенность
            # TODO: обработать по правилам поля (например, url до домена)
            if k.count('__') == 1:
                f_key, f_value = k.split('__')
                if not f_key in fkeys:
                    continue
                if k in cached_fk:
                    fk_obj = cached_fk[k]
                else:
                    fk_model = fkeys[f_key]
                    fk_obj = fk_model.objects.filter(**{f_value:v}).first()
                    if not fk_obj:
                        fk_obj = fk_model.objects.create(**{f_value:v})
                    cached_fk[k] = fk_obj
                cond.add(Q(**{f_key:fk_obj}), Q.AND)
                item[f_key] = fk_obj

        if fk_updates:
            item.update(fk_updates)

        for cond_field in cond_fields:
            cond.add(Q(**{cond_field: item.get(cond_field)}), Q.AND)
        print(cond)
        analog = model.objects.filter(cond).only('id').first()

        if analog:
            update_tasks.append({'id': analog.id, 'data': item})
        else:
            create_tasks.append(item)

    with transaction.atomic():
        for update_task in update_tasks:
            model.objects.filter(pk=update_task['id']).update(**update_task['data'])
        for create_task in create_tasks:
            model.objects.create(**create_task)


    result['created'] = len(create_tasks)
    result['updated'] = len(update_tasks)
    return result

def drop_from_excel(model, data, cond_fields: list = None):
    """Удаление массовм из базы
       :param model: модель, откуда удаляем данные
       :param data: массив с данными для удаления
       :param cond_fields: поля по которым ищем аналоги
    """
    result = {'errors': []}
    if not data:
        result['errors'].append('Не переданы данные для сохранения')
        return result
    if not cond_fields:
        result['errors'].append('Не переданы условия для обновления')
        return result
    not_found = 0
    dropped = 0
    for item in data:
        cond = Q()
        for k, v in item.items():
            if v:
                item[k] = v.strip()
        for cond_field in cond_fields:
            cond.add(Q(**{cond_field: item.get(cond_field)}), Q.AND)
        analog = model.objects.filter(cond).first()
        if analog:
            dropped += 1
            analog.delete()
        else:
            not_found += 1
    result['not_found'] = not_found
    result['dropped'] = dropped
    return result