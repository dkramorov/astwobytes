# -*- coding:utf-8 -*-
import os
import re
import time
import json
import logging
import xlsxwriter
from lxml import html as lxml_html
from io import BytesIO
from openpyxl import load_workbook

from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.urls import reverse, reverse_lazy, resolve
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.db.models import Count, Q

from apps.main_functions.files import (
    make_folder,
    ListDir,
    isForD,
    catch_file,
    unzip,
    docx2html,
    full_path,
    open_file,
    drop_file,
    copy_file,
)
from apps.main_functions.date_time import str_to_date

from apps.weld.enums import (
    JOIN_TYPES,
    WELDING_TYPES,
    MATERIALS,
    DIAMETERS,
    replace_eng2rus,
    replace_rus2eng,
)
from apps.weld.views import CUR_APP
from apps.weld.company_model import Titul, Line, Joint
from apps.weld.welder_model import Welder
from apps.weld.models import JointWelder

logger = logging.getLogger('main')

rega_digit = re.compile('[^0-9\.]+', re.I+re.U+re.DOTALL)
rega_stigma = re.compile('[0-9a-z\.]{3,6}', re.I+re.U+re.DOTALL)
rega_diameter = re.compile('[0-9]+x', re.I+re.U+re.DOTALL)
rega_thickness = re.compile('x[0-9]+', re.I+re.U+re.DOTALL)
# штуцер
rega_choke = re.compile('[Шш]туцер[ ]*([0-9,\.]{1,5})-([0-9,\.]{1,5})', re.I+re.U+re.DOTALL)
rega_sizes = re.compile('[0-9,\.]{1,5}x[0-9,\.]{1,5}', re.I+re.U+re.DOTALL)

jtypes = [join_type[1] for join_type in JOIN_TYPES]
diameters = [str(diameter) for diameter in DIAMETERS]
conn_views = [conn_view[1] for conn_view in Joint.welding_conn_view_choices]

MAX_ROWS_PREVIEW = 250
TITUL_REQUIRED_FIELDS = {
    'line': 'Номер линии',
    'joint': 'Номер узла',
    'material': 'Материал',
    'diameter': 'Диаметр, мм',
    'side_thickness': 'Толщина, мм',
    'welding_date': 'Дата сварки',
    'welding_type': 'Тип сварки',
    'welding_conn_view': 'Вид соединения',
    'join_type_from': 'Свар. эл. 1',
    'join_type_to': 'Свар. эл. 2',
    'stigma': 'Клеймо',
}

def search_problem(content, search_str):
    """Поиск чего-либо в тексте"""
    if search_str and search_str in content:
        return True

def conn_view_finder(conn_view, parts):
    """Поиск вида сварного соединения
       :param conn_view: найденны вид соединения
       :param parts: список параграфов из ячейки, где может быть правильный вид
    """
    for item in conn_views:
        if item in conn_view:
            return item
    text = ' '.join(parts)
    text = text.replace(' ', '')
    for item in conn_views:
        if item in text:
            return item

def search_stigma(text):
    text = text.replace('.', ' ')
    for item in text.split(' '):
        item = item.strip()
        if rega_stigma.search(item):
            return item
    return ''

def str_till_digit(element):
    """Встречаем число - останавливаемся
       :param element: текст, например, Штуцер25
    """
    element = element.strip()
    for j, letter in enumerate(element):
        if letter.isdigit() or letter in (' ', '-'):
            return element[:j]
    return element

def search_element_in_types(data):
    """Поиск элемента в типах соединений"""
    for part in data.lower().replace('-', ' ').split():
        for ptype in jtypes:
            if ptype in part:
                return ptype

def calc_elements(arr):
    """Подсчитываем кол-во повторений элементов в массиве
    """
    result = {}
    for item in arr:
        if not item in result:
            result[item] = 0
        result[item] += 1
    return sorted(result.items(), key=lambda item: item[1])

def search_elements(data):
    """Преобразуем текст в 2 элемента
       размер берем тот, где нету двойного размера (через слеш)
       demo_data = (
           ('Труба 57х5 –', 'Тройник П 57х5'),
           ('Тройник П 57х5 –', 'Переход ПК 57х5-32х5'),
           ('Переход ПК 57х5-32х4,5 –', 'Труба 32х4,5'),
       )
       :param data: массив элементов, например,
                    ('Труба 57х5 –', 'Тройник П 57х5')

       >>> ord(' ') # 32 - space, split by .split(' ')
       >>> ord(' ') # 160 - non breaking space, split by .split()

       проблема => Тройник П 108х6 –\nПереход П К 108х6/8-57х5
    """
    elements = {}
    for i, el in enumerate(data):
        if not el:
            continue
        el = el.strip()
        if not el:
            continue

        xcount = el.replace('х', 'x').count('x')
        if not xcount == 1:
            if xcount == 0:
                elements[i] = str_till_digit(el)
            else:
                elements[i] = search_element_in_types(el)
            eparts = el.split('-')
            for epart in eparts:
                for alt_size in epart.replace('х', 'x').replace('–', ' ').split():
                    if alt_size.count('x') == 1 and not '/' in alt_size:
                        diameter = alt_size.split('x')[0]
                        if not diameter in diameters:
                            continue
                        elements['alt_size'] = alt_size
            #continue # Брался первый элемент, хотя интересует последний

        element = el.strip()
        element = element.split()[0]
        elements[i] = str_till_digit(element)
        sizes = el.split(' ')

        for size in sizes:
            size_parts = size.replace('х', 'x').replace('-', ' ').split()
            for size_part in size_parts:
                if size_part.count('x') == 1 and not '/' in size_part:
                    diameter = size_part.split('x')[0]
                    if not diameter in diameters:
                        continue
                    elements['size'] = size_part

    if not elements.get(1) and len(data) == 1:
        z = 0
        ldata = data[0].lower()
        for part in ldata.replace('-', ' ').split():
            for ptype in jtypes:
                if ptype in part:
                    elements[z] = ptype
                    z += 1
                    break

    # Дополнительные фиксы на размер
    if elements.get(0) and elements.get(1):
        pnp = ('тройник', 'переход')
        all_data = ' '.join(data)

        # Определяем что во что переходит
        #ind_el1 = all_data.index(elements[0])
        #ind_el2 = all_data.index(elements[1])
        #if ind_el2 < ind_el1:
        #    el1 = elements[0]
        #    el2 = elements[1]
        #    elements.update({
        #        0: el2,
        #        1: el1,
        #    })

        all_data = all_data.replace('х', 'x')

        if elements[0].lower() in pnp or elements[1].lower() in pnp:
            tsize = rega_thickness.findall(all_data)
            dsize = rega_diameter.findall(all_data)

            if tsize and dsize:
                # Повторы по диаметру => реальный диаметр
                darr = calc_elements(dsize)
                diameter = darr[-1]
                # Коррекция толщины
                # Толщина по диаметру
                thickness_arr = []

                for i, item in enumerate(dsize):
                    if diameter[0] == item:
                        try:
                            thickness_arr.append(tsize[i])
                        except Exception as e:
                            logger.info('[ERROR]: %s' % e)

                if thickness_arr:
                    thickness = min([int(t.replace('x', '')) for t in thickness_arr])
                else:
                    thickness = min([int(t.replace('x', '')) for t in tsize])

                if diameter and thickness:
                    elements['size'] = '%s%s' % (diameter[0], thickness)
        elif rega_choke.search(all_data):
            choke = rega_choke.search(all_data)
            elements['size'] = '%sx%s' % (choke[1].replace(',', '.'), choke[2].replace(',', '.'))
        elif not elements.get('size') and not elements.get('alt_size'):
            alt_size = rega_sizes.search(all_data)
            if alt_size:
                elements['size'] = alt_size[0]
    return elements

def titul_parse_html_helper(tr, line_str:str, name: str, el: str = 'td'):
    """Вспомогательная функция для анализа строки таблицы
       Парсинг хтмл tr (lxml)
       :param table: таблица lxml
       :param line_str: номер линии
       :param name: название файла
       :param el: td/th
    """
    joint = {'line': line_str, 'fname': name, 'stigma': ''}
    tds = tr.xpath('.//%s' % el)
    number_joint = tds[1].xpath('.//p')
    # Пустые строки не пишем
    if not number_joint:
        return
    stigma = tds[2].xpath('.//p')
    material = tds[3].xpath('.//p')
    size = tds[4].xpath('.//p')
    date = tds[5].xpath('.//p')
    welding_type = tds[6].xpath('.//p')

    # Допишем в эксель и неразобранные данные
    raw_fields = (
        ('raw_number_joint', number_joint),
        ('raw_stigma', stigma),
        ('raw_material', material),
        ('raw_size', size),
        ('raw_date', date),
        ('raw_welding_type', welding_type),
    )
    for raw_field in raw_fields:
        value = ' '.join([item.text for item in raw_field[1] if item.text])
        joint[raw_field[0]] = value

    if number_joint:
        parts = number_joint[0].text.split(' ')
        if len(parts) == 1: # Если 1 элемент - значит, пробел забыли там
            parts = parts[0].split('.')
        joint['number'] = parts[0]
        conn_view = parts[1]
        joint['conn_view'] = conn_view_finder(conn_view, parts)
    if stigma:
        for part in stigma:
            stigma = part.text
            if not stigma:
                continue
            stigma = replace_rus2eng(stigma, force_return=True)
            s = search_stigma(stigma)
            if s:
                joint['stigma'] += ' ' + s
    if material:
        joint['material'] = material[0].text
    if size:
        data = search_elements([item.text.replace('\n', ' ') for item in size if item.text])
        joint.update(data)
    if date:
        joint['date'] = date[0].text
    if welding_type:
        joint['welding_type'] = welding_type[0].text.strip()
    return joint

def titul_parse_html(table, line_str, name):
    """Парсинг хтмл таблицы (lxml)
       № п/п (0), № стыка(1), Клеймо(2), Материал(3),
       Диаметр,мм(4), Толщина,мм(4), Дата(5)
       :param table: таблица lxml
       :param line_str: номер линии
       :param name: название файла
    """
    result = []
    trs = table.xpath('.//tbody/tr')
    # Уебаны в header таблицы умудряются засунуть стык
    trs_ths = table.xpath('.//thead/tr')
    if trs_ths:
        for tr in trs_ths:
            ths = tr.xpath('.//th')
            if ths:
                number_joint = ths[1].xpath('.//p')
                if not number_joint:
                    continue
                for par in number_joint:
                    if par.text:
                        if 'Ст' in par.text:
                            joint = titul_parse_html_helper(tr, line_str, name, 'th')
                            if joint:
                                result.append(joint)
                            break

    for tr in trs:
        joint = titul_parse_html_helper(tr, line_str, name)
        if joint:
            result.append(joint)
    return result

def titul_docx2excel(html_path: str):
    """Запихиваем все в эксельку
       :param html_path: путь до папки с html файлами
    """
    dest_fname = 'result.xlsx'
    html_files = ListDir(html_path)
    result = []
    rows = []
    for html_file in html_files:
        cur_html = os.path.join(html_path, html_file)
        if not html_file.endswith('.html'):
            drop_file(cur_html)
            continue
        content = ''
        with open_file(cur_html, 'r') as f:
            content = f.read()

        # Поиск проблем
        if search_problem(content, ''):
            print(cur_html)
        else:
            drop_file(cur_html)

        tree = lxml_html.fromstring(content)
        tables = tree.xpath('//table')
        line_str = None
        # Поиск линии
        for table in tables:
            search_line = table.xpath('.//tr/td')
            for item in search_line:
                par = item.xpath('.//p')
                if not par or not par[0].text:
                    continue
                if 'Линия' in par[0].text:
                    if len(par) == 1:
                        line_str = par[0].text.replace('Линия', '').strip()
                    else:
                        line_str = par[1].text
                    break
        for table in tables:
            ths = table.xpath('.//thead/tr/th/p')
            if ths:
                th = ths[0]
                if '№ п/п' in th.text:
                    name = html_file.replace('.html', '')
                    if '___' in name:
                        name = name.split('___')[1]
                    result += titul_parse_html(table, line_str, name)

    dest = full_path(os.path.join(html_path, dest_fname))
    book = xlsxwriter.Workbook(dest)
    sheet = book.add_worksheet('Лист 1')
    row_number = 0

    titles = (
        ('Файл', ''),
        ('Линия', 'line'),
        ('№ стыка', 'joint'),
        ('Материал', 'material'),
        ('Диаметр,мм', 'diameter'),
        ('Толщина,мм', 'side_thickness'),
        ('Дата', 'welding_date'),
        ('Тип сварки', 'welding_type'),
        ('Вид соединения', 'welding_conn_view'),
        ('Свар элемент 1', 'join_type_from'),
        ('Свар элемент 2', 'join_type_to'),
        ('Клеймо', 'stigma'),
        # Фуфел из вордовского файла баз разбора
        ('Без анализа - Стык', ''),
        ('Без анализа - Клеймо', ''),
        ('Без анализа - Материал', ''),
        ('Без анализа - Размер', ''),
        ('Без анализа - Дата', ''),
        ('Без анализа - Тип сварки', ''),
    )
    for i, title in enumerate(titles):
        sheet.write(row_number, i, title[0])
        sheet.write(row_number + 1, i, title[1])
    row_number += 1

    result_len = len(result)
    for item in result:
        row_number += 1
        number = item.get('number')
        if not number:
            continue
        if number.startswith('Ст'):
            number = number[2:]
        number = number.replace('.', '')

        diameter = side_thickness = ''
        size = item.get('size')
        if size:
            size = size.replace(',', '.')
            diameter, side_thickness = size.split('x')
            diameter = rega_digit.sub('', diameter)
            side_thickness = rega_digit.sub('', side_thickness)
        if not diameter:
            size = item.get('alt_size')
            if size:
                size = size.replace(',', '.')
                diameter, side_thickness = size.split('x')
                diameter = rega_digit.sub('', diameter)
                side_thickness = rega_digit.sub('', side_thickness)

        welding_date = item.get('date')
        if welding_date:
            # Вставляем год правильно
            welding_date = welding_date.strip()
            if welding_date and len(welding_date.strip()) == 8:
                welding_date = '%s20%s' % (welding_date[:6], welding_date[6:])

        stigma = item.get('stigma')
        try:
            stigma = replace_rus2eng(stigma)
        except Exception:
            pass

        row = (
            item.get('fname'),
            item.get('line'),
            number,
            item.get('material'),
            diameter,
            side_thickness,
            welding_date,
            item.get('welding_type'),
            item.get('conn_view'),
            item.get(0),
            item.get(1),
            stigma,
            # Фуфел из вордовского файла без разбора
            item.get('raw_number_joint'),
            item.get('raw_stigma'),
            item.get('raw_material'),
            item.get('raw_size'),
            item.get('raw_date'),
            item.get('raw_welding_type'),
        )
        for i, field in enumerate(row):
            sheet.write(row_number, i, field)

    book.close()

    result = {'errors': []}
    if result_len < MAX_ROWS_PREVIEW:
        welders = Welder.objects.all().only('id', 'stigma')
        stigmas = {welder.stigma: welder for welder in welders}
        result = prepare_data_from_excel(
            full_path(os.path.join(html_path, dest_fname)),
            stigmas=stigmas,
        )
    result['result'] = os.path.join('/media/', html_path, dest_fname)
    return result

scripts_vars = {
    'singular_obj': 'Скрипт',
    'plural_obj': 'Скрипты',
    'rp_singular_obj': 'скрипта',
    'rp_plural_obj': 'скриптов',
    'template_prefix': 'scripts/script_',
    'action_create': 'Создание',
    'action_edit': 'Редактирование',
    'action_drop': 'Удаление',
    'menu': 'scripts',
    'submenu': 'titul',
    'show_urla': 'script_titul',
}

@login_required
def script_titul(request, *args, **kwargs):
    """Анализ архива с титулом (список узлов в титуле)"""
    result = {}
    context = scripts_vars.copy()

    context['material_choices'] = MATERIALS
    context['welding_types'] = WELDING_TYPES
    context['join_types'] = JOIN_TYPES
    context['conn_view_choices'] = Joint.welding_conn_view_choices

    folder = 'scripts'
    if request.method == 'GET':
        make_folder(folder)
        # тут лежат архивы
        # нас интересуют только zip
        dirs = ListDir(folder)

    if request.is_ajax():
        return JsonResponse(result, safe=False)
    if request.FILES.get('path'):
        docx_files = None
        fname = request.FILES['path']
        if not fname.name.endswith('.zip') and not fname.name.endswith('.docx'):
            result['error'] = 'Поддерживается только zip-архив или docx'
        else:
            path = os.path.join(folder, fname.name)
            if catch_file(fname, path):
                result['success'] = 'Файл загружен'
                now = '%s' % time.time()
                now = now.replace('.', '-')
                dest_folder = os.path.join(folder, now)
                if fname.name.endswith('.zip'):
                    unzip(path, dest_folder)
                else:
                    make_folder(dest_folder)
                    copy_file(path, dest_folder)
                docx2html(dest_folder)

                preview = titul_docx2excel(dest_folder)
                result.update(preview)
                drop_file(path)

        return JsonResponse(result, safe=False)
    template = '%stitul.html' % (context['template_prefix'], )
    return render(request, template, context)

def search_welding_type(text):
    """Возвращем тип сварки
       :param text: тип сварки текстом
    """
    if not text:
        return
    text = text.upper()
    # Нестандартная проверка на комбинированный тип
    if WELDING_TYPES[0][1] in text and WELDING_TYPES[1][1] in text:
        return WELDING_TYPES[2][0]

    for welding_type in WELDING_TYPES:
        if welding_type[1] in text:
            return welding_type[0]

def search_material(text):
    """Возвращем материал
       :param text: материал текстом
    """
    for material in MATERIALS:
        if text and material[1] in text.replace('C', 'С'):
            return material[0]

def search_conn_view(text):
    """Возвращем вид соединения
       :param text: вид соединения текстом
    """
    for conn_view in Joint.welding_conn_view_choices:
        if text and conn_view[1] in text.upper():
            return conn_view[0]

def search_join_type(text):
    """Возвращем свариваемый элемент
       :param text: свариваемый элемент текстом
    """
    for join_type in JOIN_TYPES:
        if text and join_type[1] in text.lower():
            return join_type[0]

def prepare_data_from_excel(excel_file, stigmas):
    """Подготовить проанализированные данные в том виде,
       в котором они будут выведены в базу
       :param excel_file: BytesIO(excel_file.read()) для request.FILES['file']
                          или путь к файлу
       :param stigmas: словарь с клеймами
    """
    result = {'errors': []}
    wb = load_workbook(excel_file)
    sheet = wb.active

    rows = list(sheet.rows)
    if not len(rows) > 0:
        result['errors'].append('Файл пустой')
        return result

    i = 0
    errors = []
    names = {}
    data = []
    unique_joints = []
    required_keys = TITUL_REQUIRED_FIELDS.keys()
    for row in rows:
        if not names:
            for i, cell in enumerate(row):
                if cell.value in required_keys:
                    names[i] = cell.value
            if not names:
                continue

            if not len(names.keys()) == len(required_keys):
                values = names.values()
                absent_names =[name for name in required_keys if not name in values]
                result['errors'].append('обязательные поля %s' % absent_names)
                break
        item = {}
        for ind, name in names.items():
            value = row[ind].value
            if names[ind] == 'welding_date':
                value = str_to_date('%s' % value)
                if value:
                    value = value.strftime('%Y-%m-%d')
            elif names[ind] in ('join_type_from', 'join_type_to'):
                value = search_join_type(value)
            elif names[ind] == 'welding_type':
                value = search_welding_type(value)
            elif names[ind] == 'material':
                value = search_material(value)
            elif names[ind] == 'welding_conn_view':
                value = search_conn_view(value)
            elif names[ind] == 'stigma':
                values = str(value).split() if value else []
                value = ''
                for stigma in values:
                    stigma = stigma.strip()
                    if stigma in stigmas:
                        value += '%s ' % stigma
            item[name] = value
        uniq_name = '%s%s' % (item['line'], item['joint'])
        if uniq_name in unique_joints:
            logger.info('duplet %s' % uniq_name)
            continue
        unique_joints.append(uniq_name)
        data.append(item)
    result['resp'] = {'data': data[1:]}
    if result['errors']:
        result['error'] = 'Ошибка при анализе файла'
    else:
        result['success'] = 'Проверьте данные и нажмите кнопку сохранения'
    return result

@login_required
def script_titul_import(request, *args, **kwargs):
    """Загрузка титула через эксель файл"""
    result = {}
    context = scripts_vars.copy()
    context['submenu'] = 'titul_import'
    context['import_xlsx_url'] = reverse('%s:%s' % (CUR_APP, '%s_import' % scripts_vars['show_urla']))
    context['material_choices'] = MATERIALS
    context['welding_types'] = WELDING_TYPES
    context['join_types'] = JOIN_TYPES
    context['conn_view_choices'] = Joint.welding_conn_view_choices
    context['required_fields'] = TITUL_REQUIRED_FIELDS

    method = request.GET if request.method == 'GET' else request.POST
    action = method.get('action')

    welders = []
    stigmas = {}
    if action in ('import_xlsx', 'save'):
        welders = Welder.objects.all().only('id', 'stigma')
        stigmas = {welder.stigma: welder for welder in welders}

    # Загрузка только под суперадминистратором
    if action == 'save' and request.user.is_superuser:
        titul_id = request.GET.get('titul_id')
        titul = Titul.objects.filter(pk=titul_id).first()
        if not titul:
            result['error'] = 'Титул не найден'
        elif titul.line_set.all().aggregate(Count('id'))['id__count'] > 0:
            result['error'] = 'В титул уже загружены линии'
        else:
            absent_lines = []
            lines = {}
            count = int(method.get('count'))
            for i in range(count):
                joint = {}
                joint_welders = []
                for key in TITUL_REQUIRED_FIELDS.keys():
                    field = 'data[%s][%s]' % (i, key)
                    value = method.get(field)
                    if value:
                        value = value.strip()
                    if key == 'joint':
                        joint['name'] = value
                        continue
                    elif key == 'stigma':
                        values = value.split(' ')
                        for item in values:
                            if item in stigmas:
                                joint_welders.append(stigmas[item])
                        continue
                    if not value:
                        continue
                    joint[key] = value
                if not joint['line'] or len(joint['line']) < 5:
                    continue
                if not joint['line'] in lines:
                    line = Line.objects.filter(name=joint['line'], titul=titul).first()
                    if not line:
                        line = Line.objects.create(name=joint['line'], titul=titul)
                    lines[joint['line']] = line
                line = lines[joint['line']]
                joint['line'] = line
                new_joint = Joint.objects.create(**joint)
                for i, welder in enumerate(joint_welders):
                    JointWelder.objects.create(
                        joint=new_joint,
                        welder=welder,
                        position=i+1,
                    )
            result['success'] = 'Загрузка прошла успешно'

    elif action == 'import_xlsx' and request.FILES.get('file'):
        result = prepare_data_from_excel(
            BytesIO(request.FILES['file'].read()),
            stigmas=stigmas,
        )
        return JsonResponse(result, safe=False)

    if request.is_ajax():
        return JsonResponse(result, safe=False)

    template = '%stitul_upload.html' % (context['template_prefix'], )
    return render(request, template, context)
