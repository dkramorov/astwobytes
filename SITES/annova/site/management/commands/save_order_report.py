#-*- coding:utf-8 -*-
import time
import logging
import xlsxwriter
from openpyxl import load_workbook
from io import BytesIO

from django.core.management.base import BaseCommand

from apps.main_functions.files import full_path
from apps.main_functions.api_helper import open_wb

from apps.site.polis.models import Polis

logger = logging.getLogger(__name__)

"""Формирование отчета по заказам
"""

class Command(BaseCommand):
    def add_arguments(self, parser):
        # Named (optional) arguments
        parser.add_argument('--rebuild_catalogue',
            action = 'store_true',
            dest = 'rebuild_catalogue',
            default = False,
            help = 'Drop catalogue structure and recreate it')
        parser.add_argument('--cat_id',
            action = 'store',
            dest = 'cat_id',
            type = str,
            default = False,
            help = 'Set cat tag for update')

    def handle(self, *args, **options):
        started = time.time()
        polises = Polis.objects.all()

        sheet_copy = []
        wb = open_wb('report_template.xlsx')
        sheet = wb.active
        rows = sheet.rows
        for x in range(1,12):
            for y in range(1,20):
                newy = y
                if y >= 10:
                    newy = y + len(polises)
                sheet_copy.append([x, newy, sheet.cell(row=x, column=y).value])
        #print(sheet_copy)

        dest = 'report.xlsx'
        book = xlsxwriter.Workbook(full_path(dest))
        sheet = book.add_worksheet('Лист 1')
        row_number = 0

        for item in sheet_copy:
            sheet.write(item[0], item[1], item[2])

        book.close()

        elapsed = time.time() - started
        print('%.2f' % elapsed)

